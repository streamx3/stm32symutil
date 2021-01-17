#!/usr/bin/env python3
# -*- coding: utf-8 -*-

__author__ = "Andrii Shelestov"
__copyright__ = "Copyright 2019, OpenSource"
__license__ = "MIT"
__version__ = "0.1"
__email__ = "streamx3@gmail.com"

import os
import sys
import getopt
import platform
try:
    from openpyxl import Workbook
except Exception:
    print('ERROR: openpyxl package not installed.\nInstall it the way you see fit and retry.')
    sys.exit(-1)
from xml.etree import ElementTree as ET


PATH = ''

PATH_CUBEMX = ''
PATH_CUBEIDE = ''

PATH_CUBEMX_MAC_2 = '/Applications/STM32CubeMX.app/'
PATH_CUBEMX_MAC_1 = '/Applications/STMicroelectronics/STM32CubeMX.app/'
PATH_CUBEIDE_MAC = '/Applications/STM32CubeIDE.app'

PATH_CUBEMX_WIN = 'C:\\Program Files\\STMicroelectronics\\STM32Cube\\STM32CubeMX'
PATH_CUBEIDE_WIN = 'C:\\ST\\'

systems_fixedpath = ['Darwin', 'Windows']
systems_supported = ['Darwin', 'Windows', 'Linux']

URL_GITHUB = 'https://github.com/streamx3/stm32symutil'

alt_func = ['all', 'none', 'first']
alt_func_list = ['adc', 'cec', 'dac', 'dcmi', 'ddr', 'debug', 'dfsdm', 'dsihost', 'eth', 'event', 'fdcan', 'fmc',
                 'fsmc', 'hdp', 'i2c', 'jtag', 'lcd', 'ltdc', 'quad', 'rcc', 'sai', 'sdmmc', 'spdif', 'spi', 'swd',
                 'sys', 'tamp', 'tim', 'trace', 'tsc', 'uart', 'usart', 'usb', 'wkup']

input_opts = ['mcu=', 'af=', 'af_split=', 'power_split=', 'outfile=', 'help', 'version', 'mxpath=', 'idepath=']
usage_opts = ' --mcu MCU_name --af Alt_Func\n\t--power=Power_Align--outfile=OutFile\nDetails:\n' + \
             '\t--mcu [STM32F103C8Tx/STM32L552MEYxQ] -- you can type in both upper and\n\t\tlower case, ' + \
                'ommit x and everything that goes after that,\n\t\tbut onlyif anything after "x" is not ambiguous' +\
                '\n\t\t [STM32] can be ommited, e.g.: F103C8' +\
                '\n\t\t\'.\' can be added for to state the name is finite' +\
                '\n\t\t e.g.: use G071GBU6. to avoid picking G071GBU6N' +\
                '\n' + \
             '\t--af [all/none/first] or a list like adc,cec,dac,dcmi,ddr,debug,dfsdm,\n' \
             '\t\tdsihost,eth,event,fdcan,fmc,fsmc,hdp,i2c,jtag,lcd,ltdc,quad,\n' \
             '\t\trcc,sai,sdmmc,spdif,spi,swd,sys,tamp,tim,trace,tsc,tsc,uart,\n' \
             '\t\tusart,usb,wkup\n' +\
             '\t--af_split [y/n/yes/no] -- splits selected alternative functions\n\t\tin different columns\n' +\
             '\t--power_split [y/n/yes/no] -- if split selected, power pins will be put\n\t\tseparately below,' +\
             ' so they could be put into separate symbol\n' +\
             '\t--outfile [FILE] -- output filename for spreadsheet. ' +\
             '\n\t\tIf it\'s not ending with *.xlsx -- it\'ll be added for you.\n' +\
             '\t--help or -h -- display this help\n' +\
             '\t--version -- displays program\'s version\n' +\
             '\t--mxpath [PATH] -- temporal path to CubeMX on Linux\n' \
             '\t--idepath [PATH] -- temporal path to CubeIDE on Linux\n\n'\
             'Example:\n'\
             '\t --mcu stm32f103cbt --power_split y -o 1.xlsx --af spi,i2c,uart,wkup,sys'

prefix_stm32 = 'stm32'

MCU = ''
MCU_EXACT = False  # Workaround for extra letters in the MCU name, eg G071GBU6 -- G071GBU6N
AF = 'first'
AF_LIST = []
AF_SPLIT = False
POWER_SPLIT = True
OUTFILE = ''


def usage():
    print(sys.argv[0] + usage_opts)
    sys.exit(2)


def check_known_files_presence(path):
    failures = 0
    known_files = ['compatibility.xml', 'defines', 'families.xml', 'rules.xml', 'STM32F103C(8-B)Tx.xml']
    for file in known_files:
        tmp = path + os.path.sep + file
        if os.path.exists(tmp) and os.path.isfile(tmp):
            pass
        else:
            failures += 1

    if failures == 0:
        return True
    else:
        return False


def locate_internal_mx(path):
    if os.path.exists(path) and os.path.isdir(path):
        content = os.listdir(path)
        for c in content:
            if c.startswith('com.st.stm32cube.common.mx_'):
                return c
    return None


def locate_ide_win(path):
    if os.path.exists(path) and os.path.isdir(path):
        content = os.listdir(path)
        for c in content:
            if c.startswith('STM32CubeIDE_'):
                return path + c + os.path.sep + 'STM32CubeIDE' + os.path.sep
    return None


def mcufile2names(fname):
    if '(' in fname and ')' in fname:
        arr = fname.split('(')
        prefix = arr[0]
        tmp = arr[1].split(')')
        vars = tmp[0]
        suffix = tmp[1]
        vars = vars.split('-')
        rv = []
        for v in vars:
            rv.append(prefix + v + suffix)
        # print(rv)
        return rv
    else:
        return [fname]


def namecmp(found, given, exact_length=False):
    l_found = len(found)
    l_given = len(given)
    if exact_length and l_found != l_given:
        return False
    if l_found < l_given:
        return False
    for i in range(l_given):
        if found[i] != 'x':
            if found[i].lower() != given[i].lower():
                return False
    return True


if __name__ == '__main__':
    # Empty opt's check
    if len(sys.argv) < 2:
        usage()
        sys.exit()

    # Parse options:
    try:
        optlist, arglist = getopt.getopt(sys.argv[1:], 'hvo:', input_opts)
    except getopt.GetoptError as err:
        print(err)
        usage()
    for o, a in optlist:
        # print('o: ' + o + '\ta: ' + a)
        if '--help' == o or '-h' == o:
            usage()
        if '--mcu' == o:
            MCU = a
            if MCU[-1:] == '.':
                MCU_EXACT = True
                MCU = MCU[:-1]
            if MCU[:5].lower() != prefix_stm32:
                MCU = prefix_stm32 + MCU
        if '--af' == o:
            if a in alt_func:
                AF = a
            else:
                # print(a)
                final_af_list = []
                tmp_list = a.split(',')
                # print(tmp_list)
                for tmp in tmp_list:
                    if tmp in alt_func_list:
                        final_af_list.append(tmp)
                if len(final_af_list) > 0:
                    AF = 'list'
                    AF_LIST = final_af_list
                print('Validated Alternative Function options:')
                print(AF_LIST)
        if '--af_split' == o:
            if a.lower() in ['y', 'yes', 'yeah']:
                AF_SPLIT = True
        if '--power_split' == o:
            if a.lower() in ['y', 'yes', 'yeah']:
                POWER_SPLIT = True
        if '--outfile' == o or '-o' == o:
            if not a.endswith('.xlsx'):
                a += '.xlsx'
            OUTFILE = a
            print('OUTFILE: ' + OUTFILE)
        if '--version' == o or '-v' == o:
            print('Version: ' + __version__)
            print('Author : ' + __author__)
            print('License: ' + __license__)
            sys.exit(0)
        if '--mxpath' == o:
            PATH_CUBEMX = a
        if '--idepath' == o:
            PATH_CUBEIDE = a
    if OUTFILE == '':
        OUTFILE = MCU + '_' + ('-'.join(AF_LIST) if len(AF_LIST) > 0 else 'afnone') + \
                  '_' + ('spl' if AF_SPLIT else 'unspl') + \
                  '_' + ('pspl' if POWER_SPLIT else 'psunpl')
        OUTFILE += '.xlsx'
        print('You did not provide --outfile, using autogenerated: ' + OUTFILE)

    # OS detection
    system = platform.system()
    print('Dectected OS: ' + system)
    if system in systems_fixedpath:
        if system == 'Darwin':
            if os.path.exists(PATH_CUBEMX_MAC_1):
                PATH_CUBEMX = PATH_CUBEMX_MAC_1
            else:
                PATH_CUBEMX = PATH_CUBEMX_MAC_2
            PATH_CUBEIDE = PATH_CUBEIDE_MAC
        elif system == 'Windows':
            PATH_CUBEMX = PATH_CUBEMX_WIN
            PATH_CUBEIDE = PATH_CUBEIDE_WIN
    else:
        if system in systems_supported:
            if PATH == '' and PATH_CUBEMX == '' and PATH_CUBEIDE == '':
                print('You\'re using Linux, so you should open this file and set one or both of first 2 variables,',
                      'which are:\n' +
                      '\tPATH_CUBEMX - path to CubeMX, e.g:"/.../STM32CubeMX/"\n' +
                      '\tPATH_CUBEIDE - path to CubeIDE, e.g:"/.../STM32CubeIDE/"\n\n' +
                      'Restart this program after doing so. Cube MX will be preferred if both set.\n' +
                      'Alternatively use --mxpath or --idepath')
                sys.exit(-1)

    # Path verification
    if PATH == '':
        TMP = PATH_CUBEMX
        if TMP[-1:] != os.path.sep:
            TMP += os.path.sep
        if system == 'Darwin':
            TMP += 'Contents/Resources/'
        TMP += 'db' + os.path.sep + 'mcu' + os.path.sep
        if os.path.exists(PATH_CUBEMX) and os.path.isdir(PATH_CUBEMX) and os.path.exists(TMP) and os.path.isdir(TMP):
            if check_known_files_presence(TMP):
                PATH = TMP

        if PATH == '':
            if system == 'Windows':
                PATH_CUBEIDE_SP = locate_ide_win(PATH_CUBEIDE)
            else:
                PATH_CUBEIDE_SP = PATH_CUBEIDE
            if os.path.exists(PATH_CUBEIDE_SP) and os.path.isdir(PATH_CUBEIDE_SP):
                TMP = PATH_CUBEIDE_SP
                if TMP[-1:] != os.path.sep:
                    TMP += os.path.sep
                if system == 'Darwin':
                    TMP += 'Contents/Eclipse/'
                TMP += 'plugins' + os.path.sep
                internal_mx = locate_internal_mx(TMP)
                if internal_mx is None:
                    print('CubeMX not found and CubeIDE was not parsed correctly.\n' +
                          'Please install CubeMX/CubeIDE or report a bug:\n' + URL_GITHUB)
                    sys.exit(-1)
                TMP += internal_mx + os.path.sep + 'db' + os.path.sep + 'mcu' + os.path.sep
                if os.path.exists(TMP) and os.path.isdir(TMP):
                    if check_known_files_presence(TMP):
                        PATH = TMP
        if PATH == '':
            if system == 'Linux':
                print('Found nothing. Make sure you\'ve set PATH or give and installed CubeMX or CubeIDE.\n' +
                      'Or, if you suspect a bug -- feel free to report it:\n' + URL_GITHUB)
            elif system in systems_fixedpath:
                print('Found nothing. Make sure you\'ve installed CubeMX or CubeIDE.\n' +
                      'If you\'re sure CubeMX or CubeIDE is installed -- please report a bug:\n' + URL_GITHUB)
            else:
                print('Well, you are the one who uses this script in ' + system + ' environment.\n' +
                      'What is the deal with ' + system + ' anyway?\n' +
                      'Probable cause of your problen is a lack of ST\'s CUBE software or wrong a  path.\n' +
                      'But you probably know what you\'re doing, so good luck!')
            sys.exit(-1)
    print('Using PATH: ' + PATH)

    # MCU search
    print('Searching for: ' + MCU.upper())
    files_list = os.listdir(PATH)
    d_mcus = {}
    for f in files_list:
        f_red = f.replace('.xml', '')
        if f.startswith('STM32'):
            d_mcus[f] = mcufile2names(f_red)
    mcus_found = []
    for k in d_mcus:
        for tmp_mcu in d_mcus[k]:
            if namecmp(tmp_mcu, MCU, MCU_EXACT):
            # if tmp_mcu.replace('x', '').lower().startswith(MCU.lower()):
                if k not in mcus_found:
                    mcus_found.append(k)
    if len(mcus_found) == 0:
        print('MCU not found!')
        sys.exit(0)
    if len(mcus_found) > 1:
        print('Found ' + str(len(mcus_found)) + ' MCUs:')
        for m in mcus_found:
            print(m)
        print('Provide more specific input!')
        sys.exit(0)
    if len(mcus_found) == 1:
        print('Successfully found 1 MCU: ' + mcus_found[0])

    # Parsing XML
    tree = ET.parse(PATH + os.path.sep + mcus_found[0])
    root = tree.getroot()
    tag = root.tag
    # print(tag)
    # print(type(tag))
    table = {}
    table_power = {}
    table_vdd = {}
    table_vss = {}
    for ch1 in root:
        t = ch1.tag.replace('{http://mcd.rou.st.com/modules.php?name=mcu}', '')
        attrib = ch1.attrib
        # print(t, attrib)
        if t == 'Pin':
            af = []
            for ch2 in ch1:
                t = ch2.tag.replace('{http://mcd.rou.st.com/modules.php?name=mcu}', '')
                # print(t, ch2.attrib)
                att2 = ch2.attrib
                if 'Name' in att2:
                    if AF == 'list':
                        for a in AF_LIST:
                            ch2name = ch2.attrib['Name']
                            if ch2name.lower().find(a) > -1:
                                af.append(ch2.attrib['Name'])
                    if AF in ['all', 'first']:
                        af.append(ch2.attrib['Name'])
                        if AF == 'first':
                            break

            if not AF_SPLIT:
                af = ['/'.join(af)]

            pos = attrib['Position']
            del attrib['Position']
            attrib['af'] = af
            attrib['Type'] = attrib['Type'].replace('Reset', 'Input').replace('MonoIO', 'I/O')
            attrib['Type'] = attrib['Type'].replace('Boot', 'Input')
            if POWER_SPLIT:
                if attrib['Type'] == 'Power':
                    if attrib['Name'].lower() == 'vdd':
                        table_vdd[pos] = attrib
                    elif attrib['Name'].lower() == 'vss':
                        table_vss[pos] = attrib
                    else:
                        table_power[pos] = attrib
                else:
                    table[pos] = attrib
            else:
                table[pos] = attrib
    # Cheking file availability
    if os.path.exists(OUTFILE):
        try:
            myfile = open(OUTFILE, 'r+')  # or "a+", whatever you need
        except IOError:
            print('Could not open file! Please close Excel!')
            sys.exit(-1)
        myfile.close()
    # Printing table
    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'Designator'
    ws['B1'] = 'Electrical Type'
    ws['C1'] = 'Display Name'
    if AF_SPLIT:
        ws.merge_cells('D1:S1')
        ws['D1'] = 'Alternative Fucntions'

    i = 2
    alphabet = [#'A', 'B', 'C',
                'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T',
                'U', 'V', 'W', 'X', 'Y', 'Z']
    for k in table:
        ws['A' + str(i)] = k
        ws['B' + str(i)] = table[k]['Type']
        row = table[k]
        if AF_SPLIT:
            ws['C' + str(i)] = row['Name']
            if len(row['af']) > 0:
                letter = 0
                for af in row['af']:
                    ws[str(alphabet[letter]) + str(i)] = af
                    letter += 1

        else:
            string = row['Name']
            if len(row['af']) == 1 and type(row['af'][0]) is str and row['af'][0] != '':
                string += '/' + table[k]['af'][0]
            ws['C' + str(i)] = string
        i += 1
    ws['A' + str(i)] = 'Total: '
    ws['B' + str(i)] = str(len(table))
    i += 1
    if POWER_SPLIT:
        i += 1
        for t in [table_power, table_vdd, table_vss]:
            for k in t:
                ws['A' + str(i)] = k
                ws['B' + str(i)] = t[k]['Type']
                ws['C' + str(i)] = t[k]['Name']
                i += 1
        ws['A' + str(i)] = 'Total: '
        ws['B' + str(i)] = str(len(table_power) + len(table_vdd) + len(table_vss))
        i += 1
    wb.save(OUTFILE)
    print('Done!')
